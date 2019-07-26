import os
from getpathinfo import get_path
from openpyxl import load_workbook

filepath = get_path()
print(filepath)

"""
读取xlsx表格
"""


class ReadExcel(object):
    def get_xlsx(self, xlsx_name, sheet_name):
        excel_list = []
        excelpath = os.path.join(filepath, 'testCase', xlsx_name)
        excel = load_workbook(excelpath)  # 加载文件
        # 获取sheet
        sheet = excel[sheet_name]  # 获取表名
        # 获取行数和列数
        # maxrow = sheet.max_row  # 获取行数
        # maxcol = sheet.max_column  # 获取列数

        for row in sheet.rows:  # 根据行数做循环
            son_list = []
            for cell in row:
                son_list.append(cell.value)  # 将每一行的数据添加到son_list列表里面
            if son_list[1] != u'case_name':  # son_list的第2列不等于case_name那么我们把这行的数据添加到excel_list = []
                excel_list.append(son_list)
        return excel_list


if __name__ == '__main__':  # 我们执行该文件测试一下是否可以正确获取Excel中的值
    print(ReadExcel().get_xlsx('interface_usecases.xlsx', 'login'))

